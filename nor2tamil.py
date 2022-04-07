import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
from openpyxl import load_workbook
from tamil.utf8 import get_letters, get_tamil_words


driver=webdriver.Chrome()
driver.get('https://glosbe.com/nb/ta/')
 
df = pd.read_excel('eng.xlsx',engine='openpyxl')
wb= load_workbook('eng.xlsx')
# define the sheet name
ws=wb['Ark1']
# slice the dataframe with from and to indexes
word = df.iloc[5904:6010]
# excel row number    
j=2

# iterating all norwegian words one by one
for i in word['Norwegian\n']:
    drop_down = driver.find_element_by_css_selector("input[id='dictionary-search'][name='q']")
    drop_down.send_keys(str(i))
    try:
        driver.find_element_by_css_selector("button[data-element='search'][type='submit']").click()
        element = driver.find_element_by_xpath("//div[@class='phrase__translation__section']/div/ul/li/div").text
        element = get_tamil_words(get_letters(element)) 
        li = ''.join([str(e) for e in element])
        ws.cell(row=j,column=6,value=str(li))
        j+=1
        wb.save('eng.xlsx')
    except:
        j+=1
        continue

      
 


